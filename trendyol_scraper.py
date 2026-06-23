import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import re
import time
import datetime
import glob
import schedule
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ─────────────────────────────────────────────
#  SELENIUM IMPORT WITH GRACEFUL FALLBACK
# ─────────────────────────────────────────────
# The error "No module named 'selenium.webdriver.chrome.webdriver'" means
# PyInstaller did not bundle all Selenium internals. We work around this by
# importing the chrome service/options modules explicitly so PyInstaller
# can detect them, and by using webdriver-manager to supply chromedriver.

try:
    import gspread
    from google.oauth2.service_account import Credentials
    _GSPREAD_AVAILABLE = True
except ImportError:
    _GSPREAD_AVAILABLE = False

from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, WebDriverException

try:
    from webdriver_manager.chrome import ChromeDriverManager
    _WDM_AVAILABLE = True
except ImportError:
    _WDM_AVAILABLE = False

# ─────────────────────────────────────────────
#  SCRAPER LOGIC (Selenium-based)
# ─────────────────────────────────────────────

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

PAGE_LOAD_TIMEOUT = 30
ELEMENT_WAIT_TIMEOUT = 15
DELAY_BETWEEN_PRODUCTS = 1.0


def _make_chrome_options(headless: bool) -> ChromeOptions:
    options = ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1366,1000")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=ro-RO")
    options.add_argument(f"user-agent={USER_AGENT}")
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    return options


def _apply_stealth(driver):
    try:
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
        )
    except Exception:
        pass


def build_driver(headless: bool = True):
    """Create a Chrome WebDriver using the most reliable method available.

    Three-attempt strategy:
    1. webdriver-manager: downloads & caches the exact matching chromedriver.
       Most reliable when the network allows it.
    2. Selenium Manager (built-in): works on Selenium 4.6+ without extra packages.
    3. Bare webdriver.Chrome(): last resort — works if chromedriver is on PATH.
    """
    errors = []

    # ── Attempt 1: webdriver-manager (most explicit, most reliable) ──
    if _WDM_AVAILABLE:
        try:
            service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=_make_chrome_options(headless))
            driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
            _apply_stealth(driver)
            return driver
        except Exception as e:
            errors.append(f"webdriver-manager: {e}")

    # ── Attempt 2: Selenium Manager (no extra package needed) ──
    try:
        driver = webdriver.Chrome(options=_make_chrome_options(headless))
        driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
        _apply_stealth(driver)
        return driver
    except Exception as e:
        errors.append(f"Selenium Manager: {e}")

    # ── Attempt 3: chromedriver on PATH ──
    try:
        service = ChromeService()
        driver = webdriver.Chrome(service=service, options=_make_chrome_options(headless))
        driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
        _apply_stealth(driver)
        return driver
    except Exception as e:
        errors.append(f"PATH chromedriver: {e}")

    raise RuntimeError(
        "Could not start Chrome after 3 attempts.\n\n"
        + "\n".join(errors)
        + "\n\nFix: reinstall dependencies with:  pip install --upgrade --force-reinstall selenium webdriver-manager"
    )


# ─────────────────────────────────────────────
#  COOKIE BANNER
# ─────────────────────────────────────────────

_COOKIE_BUTTON_CANDIDATES = [
    (By.ID, "onetrust-accept-btn-handler"),
    (By.CSS_SELECTOR, "button#cookie-accept-all"),
    (By.XPATH, "//button[contains(., 'Kabul Et')]"),
    (By.XPATH, "//button[contains(., 'Accept')]"),
]


def _dismiss_cookie_banner(driver):
    for by, selector in _COOKIE_BUTTON_CANDIDATES:
        try:
            elems = driver.find_elements(by, selector)
            if elems and elems[0].is_displayed():
                elems[0].click()
                return True
        except Exception:
            continue
    return False


# ─────────────────────────────────────────────
#  COUNTRY SELECTION (Romania)
# ─────────────────────────────────────────────

def _dismiss_country_modal(driver):
    """Select Romania in the country selection modal if it appears."""
    try:
        select_el = WebDriverWait(driver, 6).until(
            EC.presence_of_element_located((By.ID, "country-select"))
        )
        if select_el.is_displayed():
            Select(select_el).select_by_visible_text("Romania")
            time.sleep(0.4)
            confirm_btn = driver.find_element(
                By.CSS_SELECTOR, "[data-testid='country-select-btn-desktop']"
            )
            confirm_btn.click()
            time.sleep(1.5)
            return True
    except Exception:
        pass
    return False


# ─────────────────────────────────────────────
#  GOOGLE SHEETS SYNC
# ─────────────────────────────────────────────

GSHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def sync_to_google_sheets(results: list, sheet_url: str, prev_prices: dict, credentials_path: str, log_fn=None):
    """Write/overwrite the Google Sheet with current scrape results.

    Args:
        results:          list of scrape result dicts
        sheet_url:        full Google Sheets URL shared by the user
        prev_prices:      dict of {url: previous_price}
        credentials_path: path to the service account JSON key file
        log_fn:           optional callable(msg, level) for GUI logging
    """
    def _log(msg, level="info"):
        if log_fn:
            log_fn(msg, level)

    if not _GSPREAD_AVAILABLE:
        _log("gspread not installed — skipping Google Sheets sync. Run: pip install gspread google-auth", "warn")
        return False

    if not sheet_url or not sheet_url.strip():
        return False

    if not credentials_path or not os.path.exists(credentials_path):
        _log("Google Sheets credentials file not found — skipping sync.", "warn")
        return False

    try:
        creds = Credentials.from_service_account_file(credentials_path, scopes=GSHEETS_SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_url(sheet_url.strip())
        ws = sh.sheet1

        headers = [
            "#", "Product Name", "EAN", "URL",
            "Original Price", "Discount %", "Price After Discount",
            "Old Price (Prev Day)", "Status", "Last Updated"
        ]

        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = [headers]
        for i, r in enumerate(results, 1):
            rows.append([
                i,
                r.get("product_name", ""),
                r.get("ean", ""),
                r.get("url", ""),
                r.get("original_price") or "",
                r.get("discount_pct") or "",
                r.get("discounted_price") or "",
                prev_prices.get(r["url"], ""),
                "OK" if r.get("status") == "ok" else "ERROR",
                now_str,
            ])

        ws.clear()
        ws.update(rows, value_input_option="USER_ENTERED")
        _log(f"Google Sheets updated — {len(results)} rows written.", "head")
        return True

    except Exception as e:
        _log(f"Google Sheets sync failed: {e}", "error")
        return False




def parse_product_page(html: str):
    soup = BeautifulSoup(html, "lxml")

    name_el = soup.find("h1", class_="product-title")
    name = name_el.get_text(strip=True) if name_el else ""

    sale_int = soup.find("span", {"data-testid": "integer-part"})
    sale_dec = soup.find("span", {"data-testid": "decimal-part"})
    discounted_price = None
    if sale_int:
        price_str = sale_int.get_text(strip=True).replace(",", "")
        if sale_dec:
            price_str += "." + sale_dec.get_text(strip=True).replace(",", "")
        try:
            discounted_price = float(price_str)
        except ValueError:
            pass

    orig_el = soup.find("div", {"data-testid": "strikethrough-price"})
    original_price = None
    if orig_el:
        orig_str = orig_el.get_text(strip=True).replace(",", ".").replace("Lei", "").strip()
        try:
            original_price = float(orig_str)
        except ValueError:
            pass

    discount_el = soup.find("div", {"data-testid": "discount-badge"})
    discount_pct = None
    if discount_el:
        d_text = discount_el.get_text(strip=True).replace("-", "").replace("%", "").strip()
        try:
            discount_pct = float(d_text)
        except ValueError:
            pass

    ean = ""
    barcode_match = re.search(r"Cod de bare nr\.?:?\s*([\w\d]+)", html, re.IGNORECASE)
    if barcode_match:
        ean = barcode_match.group(1).strip()

    if original_price is None and discounted_price:
        original_price = discounted_price

    if discount_pct is None and original_price and discounted_price and original_price > 0:
        discount_pct = round((original_price - discounted_price) / original_price * 100, 1)

    return {
        "product_name": name,
        "ean": ean,
        "original_price": original_price,
        "discounted_price": discounted_price if discounted_price else original_price,
        "discount_pct": discount_pct or 0,
    }


def scrape_product(driver, url: str, ean_from_excel=None):
    result = {
        "url": url,
        "product_name": "",
        "ean": ean_from_excel or "",
        "original_price": None,
        "discount_pct": None,
        "discounted_price": None,
        "status": "ok",
        "error": "",
    }

    try:
        driver.get(url)
        _dismiss_country_modal(driver)   # Romania selection if modal appears
        _dismiss_cookie_banner(driver)

        try:
            WebDriverWait(driver, ELEMENT_WAIT_TIMEOUT).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "span[data-testid='integer-part'], h1.product-title")
                )
            )
        except TimeoutException:
            pass

        time.sleep(0.6)

        data = parse_product_page(driver.page_source)

        if data and (data.get("product_name") or data.get("discounted_price")):
            result.update(data)
            if ean_from_excel and not result.get("ean"):
                result["ean"] = ean_from_excel
        else:
            result["status"] = "error"
            result["error"] = "Could not extract data (page may not have loaded)"

    except TimeoutException:
        result["status"] = "error"
        result["error"] = "Timed out loading page"
    except WebDriverException as e:
        result["status"] = "error"
        result["error"] = str(e).splitlines()[0][:150]
    except Exception as e:
        result["status"] = "error"
        result["error"] = str(e)[:150]

    return result


# ─────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────

def find_previous_excel(output_folder: str):
    files = sorted(glob.glob(os.path.join(output_folder, "trendyol_*.xlsx")))
    if files:
        return files[-1]
    return None


def load_previous_prices(prev_file: str):
    prev_prices = {}
    if not prev_file or not os.path.exists(prev_file):
        return prev_prices
    try:
        df = pd.read_excel(prev_file)
        for _, row in df.iterrows():
            url = str(row.get("URL", "")).strip()
            price = row.get("Price After Discount")
            if url and price is not None:
                try:
                    prev_prices[url] = float(price)
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass
    return prev_prices


def save_results_to_excel(results: list, output_folder: str, prev_prices: dict):
    os.makedirs(output_folder, exist_ok=True)
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(output_folder, f"trendyol_{date_str}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt_fill    = PatternFill("solid", fgColor="EBF2FA")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    ok_fill     = PatternFill("solid", fgColor="D6F4E0")
    err_fill    = PatternFill("solid", fgColor="FADADD")
    price_font  = Font(name="Arial", size=10, color="1E3A5F", bold=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al     = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        "#", "Product Name", "EAN", "URL",
        "Original Price", "Discount %", "Price After Discount",
        "Old Price (Prev Day)", "Status"
    ]
    col_widths = [5, 50, 20, 60, 16, 14, 20, 20, 10]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_al
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, r in enumerate(results, 1):
        row_num = i + 1
        fill = alt_fill if i % 2 == 0 else white_fill
        old_price = prev_prices.get(r["url"])

        values = [
            i,
            r.get("product_name", ""),
            r.get("ean", ""),
            r.get("url", ""),
            r.get("original_price"),
            r.get("discount_pct"),
            r.get("discounted_price"),
            old_price,
            "✓ OK" if r.get("status") == "ok" else "✗ ERROR",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = border
            cell.fill = fill
            cell.alignment = left_al if col_idx in (2, 4) else center_al
            cell.font = Font(name="Arial", size=10)

            if col_idx in (5, 7, 8) and val is not None:
                cell.number_format = '#,##0.00'
                cell.font = price_font
            if col_idx == 6 and val is not None:
                cell.number_format = '0.0"%"'

            if col_idx == 9:
                cell.fill = ok_fill if r.get("status") == "ok" else err_fill
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color="1A7F3C" if r.get("status") == "ok" else "CC0000")

        ws.row_dimensions[row_num].height = 20

    summary_row = len(results) + 2
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, name="Arial")
    ws.cell(row=summary_row, column=2, value=f"Total: {len(results)} products")
    ok_count = sum(1 for r in results if r.get("status") == "ok")
    ws.cell(row=summary_row, column=3, value=f"OK: {ok_count}  |  Errors: {len(results)-ok_count}")

    wb.save(filename)
    return filename


# ─────────────────────────────────────────────
#  GUI APPLICATION
# ─────────────────────────────────────────────

class TrendyolScraperApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Trendyol Price Scraper")
        self.geometry("900x680")
        self.resizable(True, True)
        self.configure(bg="#0F2640")

        self._scheduler_thread = None
        self._scheduled_time = None
        self._running = False
        self._stop_scrape = False
        self._schedule_active = False
        self._driver = None

        self._build_ui()
        self._apply_theme()

    def _build_ui(self):
        top = tk.Frame(self, bg="#0F2640", pady=12)
        top.pack(fill="x")
        tk.Label(top, text="🛒 Trendyol Price Scraper",
                 font=("Segoe UI", 20, "bold"), bg="#0F2640", fg="#F47425").pack()
        tk.Label(top, text="Automated daily price monitoring for your product list",
                 font=("Segoe UI", 10), bg="#0F2640", fg="#8EAAC8").pack()

        content = tk.Frame(self, bg="#F0F4F8")
        content.pack(fill="both", expand=True, padx=0, pady=0)

        left = tk.Frame(content, bg="#F0F4F8", padx=20, pady=16)
        left.pack(side="left", fill="y")

        right = tk.Frame(content, bg="#FFFFFF", padx=16, pady=16)
        right.pack(side="right", fill="both", expand=True)

        self._section(left, "📁 Input / Output")

        tk.Label(left, text="Excel File (URLs):", **self._lbl_style()).pack(anchor="w", pady=(6,2))
        row_xls = tk.Frame(left, bg="#F0F4F8")
        row_xls.pack(fill="x", pady=(0,8))
        self.excel_var = tk.StringVar()
        tk.Entry(row_xls, textvariable=self.excel_var, **self._entry_style(), width=28).pack(side="left", fill="x", expand=True)
        tk.Button(row_xls, text="Browse", command=self._browse_excel, **self._btn_small()).pack(side="left", padx=(6,0))

        tk.Label(left, text="Output Folder:", **self._lbl_style()).pack(anchor="w", pady=(0,2))
        row_out = tk.Frame(left, bg="#F0F4F8")
        row_out.pack(fill="x", pady=(0,12))
        self.output_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "TrendyolReports"))
        tk.Entry(row_out, textvariable=self.output_var, **self._entry_style(), width=28).pack(side="left", fill="x", expand=True)
        tk.Button(row_out, text="Browse", command=self._browse_output, **self._btn_small()).pack(side="left", padx=(6,0))

        self._section(left, "📊 Google Sheets (optional)")

        tk.Label(left, text="Sheet URL:", **self._lbl_style()).pack(anchor="w", pady=(6,2))
        self.gsheet_var = tk.StringVar()
        tk.Entry(left, textvariable=self.gsheet_var, **self._entry_style(), width=36).pack(fill="x", pady=(0,6))

        tk.Label(left, text="Service Account JSON:", **self._lbl_style()).pack(anchor="w", pady=(0,2))
        row_creds = tk.Frame(left, bg="#F0F4F8")
        row_creds.pack(fill="x", pady=(0,12))
        self.creds_var = tk.StringVar()
        tk.Entry(row_creds, textvariable=self.creds_var, **self._entry_style(), width=28).pack(side="left", fill="x", expand=True)
        tk.Button(row_creds, text="Browse", command=self._browse_creds, **self._btn_small()).pack(side="left", padx=(6,0))

        self._section(left, "🌐 Browser")
        self.headless_var = tk.BooleanVar(value=True)
        tk.Checkbutton(left, text="Run headless (hide Chrome window)",
                       variable=self.headless_var, bg="#F0F4F8", fg="#2C3E50",
                       activebackground="#F0F4F8", font=("Segoe UI", 9),
                       selectcolor="#FFFFFF").pack(anchor="w", pady=(2,8))

        self._section(left, "⏰ Schedule")
        tk.Label(left, text="Run daily at:", **self._lbl_style()).pack(anchor="w", pady=(6,2))

        time_row = tk.Frame(left, bg="#F0F4F8")
        time_row.pack(anchor="w", pady=(0,8))
        self.hour_var = tk.StringVar(value="08")
        self.min_var  = tk.StringVar(value="00")

        hour_spin = ttk.Spinbox(time_row, from_=0, to=23, width=4,
                                textvariable=self.hour_var, format="%02.0f",
                                font=("Segoe UI", 13, "bold"))
        hour_spin.pack(side="left")
        tk.Label(time_row, text=":", font=("Segoe UI", 14, "bold"),
                 bg="#F0F4F8", fg="#0F2640").pack(side="left", padx=2)
        min_spin = ttk.Spinbox(time_row, from_=0, to=59, width=4,
                               textvariable=self.min_var, format="%02.0f",
                               font=("Segoe UI", 13, "bold"))
        min_spin.pack(side="left")

        self.schedule_btn = tk.Button(left, text="▶  Activate Schedule",
                                      command=self._toggle_schedule,
                                      **self._btn_primary())
        self.schedule_btn.pack(fill="x", pady=(4, 12))

        self.schedule_lbl = tk.Label(left, text="Schedule: inactive",
                                     font=("Segoe UI", 9), bg="#F0F4F8", fg="#888")
        self.schedule_lbl.pack(anchor="w")

        self._section(left, "▶  Run Now")
        self.run_btn = tk.Button(left, text="▶  Scrape Now",
                                 command=self._run_now, **self._btn_action())
        self.run_btn.pack(fill="x", pady=(8,4))
        self.stop_btn = tk.Button(left, text="■  Stop",
                                  command=self._stop_now, state="disabled",
                                  **self._btn_stop())
        self.stop_btn.pack(fill="x", pady=(0,4))

        tk.Label(right, text="📋 Activity Log", font=("Segoe UI", 12, "bold"),
                 bg="#FFFFFF", fg="#0F2640").pack(anchor="w", pady=(0,8))

        self.log_frame = tk.Frame(right, bg="#FFFFFF")
        self.log_frame.pack(fill="both", expand=True)

        self.log_text = tk.Text(self.log_frame, font=("Consolas", 9),
                                bg="#0D1B2A", fg="#A8D8A8",
                                relief="flat", wrap="word",
                                padx=8, pady=8, state="disabled")
        scrollbar = ttk.Scrollbar(self.log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)

        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(right, variable=self.progress_var,
                                        maximum=100, style="green.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(8,0))

        self.status_lbl = tk.Label(right, text="Ready",
                                   font=("Segoe UI", 9), bg="#FFFFFF", fg="#555")
        self.status_lbl.pack(anchor="w")

    def _section(self, parent, title):
        tk.Label(parent, text=title, font=("Segoe UI", 10, "bold"),
                 bg="#F0F4F8", fg="#0F2640").pack(anchor="w", pady=(12,0))
        tk.Frame(parent, height=1, bg="#D0D8E4").pack(fill="x", pady=(2,4))

    def _apply_theme(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("green.Horizontal.TProgressbar",
                        troughcolor="#E0E8F0", background="#27AE60",
                        thickness=8)
        style.configure("TSpinbox", fieldbackground="white", padding=4)

    def _lbl_style(self):
        return {"font": ("Segoe UI", 9, "bold"), "bg": "#F0F4F8", "fg": "#2C3E50"}

    def _entry_style(self):
        return {"font": ("Segoe UI", 9), "relief": "flat",
                "highlightbackground": "#B0C4DE", "highlightthickness": 1,
                "bg": "white", "fg": "#2C3E50"}

    def _btn_small(self):
        return {"font": ("Segoe UI", 8), "bg": "#2C5F8A", "fg": "white",
                "relief": "flat", "padx": 8, "pady": 4, "cursor": "hand2",
                "activebackground": "#1A4066", "activeforeground": "white"}

    def _btn_primary(self):
        return {"font": ("Segoe UI", 10, "bold"), "bg": "#F47425", "fg": "white",
                "relief": "flat", "padx": 12, "pady": 8, "cursor": "hand2",
                "activebackground": "#C05A15", "activeforeground": "white"}

    def _btn_action(self):
        return {"font": ("Segoe UI", 10, "bold"), "bg": "#27AE60", "fg": "white",
                "relief": "flat", "padx": 12, "pady": 8, "cursor": "hand2",
                "activebackground": "#1E8449", "activeforeground": "white"}

    def _btn_stop(self):
        return {"font": ("Segoe UI", 10, "bold"), "bg": "#C0392B", "fg": "white",
                "relief": "flat", "padx": 12, "pady": 8, "cursor": "hand2",
                "activebackground": "#922B21", "activeforeground": "white"}

    def _browse_creds(self):
        f = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if f:
            self.creds_var.set(f)

    def _browse_excel(self):
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if f:
            self.excel_var.set(f)

    def _browse_output(self):
        d = filedialog.askdirectory()
        if d:
            self.output_var.set(d)

    def log(self, msg: str, level="info"):
        colors = {"info": "#A8D8A8", "warn": "#F9E79F",
                  "error": "#F1948A", "head": "#85C1E9"}
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        full = f"[{timestamp}] {msg}\n"
        self.log_text.config(state="normal")
        self.log_text.tag_config(level, foreground=colors.get(level, "#A8D8A8"))
        self.log_text.insert("end", full, level)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _set_status(self, txt):
        self.status_lbl.config(text=txt)
        self.update_idletasks()

    def _toggle_schedule(self):
        if not self._schedule_active:
            h = self.hour_var.get().zfill(2)
            m = self.min_var.get().zfill(2)
            try:
                _ = int(h); _ = int(m)
                if not (0 <= int(h) <= 23 and 0 <= int(m) <= 59):
                    raise ValueError
            except ValueError:
                messagebox.showerror("Invalid time", "Please enter a valid HH:MM time.")
                return

            self._scheduled_time = f"{h}:{m}"
            schedule.clear()
            schedule.every().day.at(self._scheduled_time).do(self._run_scrape)
            self._schedule_active = True
            self.schedule_btn.config(text="⏹  Deactivate Schedule", bg="#888")
            self.schedule_lbl.config(text=f"Schedule: active — runs daily at {h}:{m}",
                                     fg="#27AE60")
            self.log(f"Schedule set: daily at {h}:{m}", "head")

            if self._scheduler_thread is None or not self._scheduler_thread.is_alive():
                self._scheduler_thread = threading.Thread(
                    target=self._scheduler_loop, daemon=True)
                self._scheduler_thread.start()
        else:
            schedule.clear()
            self._schedule_active = False
            self.schedule_btn.config(text="▶  Activate Schedule", bg="#F47425")
            self.schedule_lbl.config(text="Schedule: inactive", fg="#888")
            self.log("Schedule deactivated.", "warn")

    def _scheduler_loop(self):
        while self._schedule_active:
            schedule.run_pending()
            time.sleep(15)

    def _run_now(self):
        if self._running:
            return
        thread = threading.Thread(target=self._run_scrape, daemon=True)
        thread.start()

    def _stop_now(self):
        self._stop_scrape = True
        self.log("Stop requested — finishing current product…", "warn")

    def _run_scrape(self):
        self._running = True
        self._stop_scrape = False
        self.run_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.progress_var.set(0)

        excel_path  = self.excel_var.get().strip()
        output_path = self.output_var.get().strip()
        sheet_url   = self.gsheet_var.get().strip()
        creds_path  = self.creds_var.get().strip()

        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("Error", "Please select a valid Excel file.")
            self._done()
            return

        # At least one output destination must be configured
        if not output_path and not sheet_url:
            messagebox.showerror(
                "No output configured",
                "Please provide at least one of:\n"
                "  • Output Folder (for Excel file)\n"
                "  • Google Sheet URL"
            )
            self._done()
            return

        self.log("══════════════════════════════════════", "head")
        self.log(f"Starting scrape — {datetime.datetime.now():%Y-%m-%d %H:%M:%S}", "head")

        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            self.log(f"Cannot read Excel: {e}", "error")
            self._done()
            return

        url_col = next((c for c in df.columns if "url" in c.lower() or "trendyol" in c.lower()), None)
        ean_col = next((c for c in df.columns if "ean" in c.lower()), None)

        if not url_col:
            self.log("No URL column found in Excel.", "error")
            self._done()
            return

        urls = df[url_col].dropna().tolist()
        eans = df[ean_col].astype(str).tolist() if ean_col else [""] * len(urls)
        total = len(urls)
        self.log(f"Loaded {total} URLs from '{os.path.basename(excel_path)}'", "info")

        prev_file = find_previous_excel(output_path) if output_path else None
        if prev_file:
            self.log(f"Previous file found: {os.path.basename(prev_file)}", "info")
        prev_prices = load_previous_prices(prev_file)

        self.log("Launching Chrome (this can take a few seconds the first time)…", "head")
        try:
            self._driver = build_driver(headless=self.headless_var.get())
        except Exception as e:
            err_msg = str(e)
            self.log(f"Could not start Chrome: {err_msg}", "error")
            self.log("─── Troubleshooting ───", "warn")
            self.log("1. Make sure Google Chrome is installed.", "warn")
            self.log("2. Run: pip install --upgrade --force-reinstall selenium webdriver-manager", "warn")
            messagebox.showerror(
                "Chrome not found",
                "Could not launch Chrome.\n\n"
                "Please check:\n"
                "1. Google Chrome is installed\n"
                "2. Run in terminal:\n"
                "   pip install --upgrade --force-reinstall selenium webdriver-manager\n\n"
                f"Error: {err_msg[:300]}"
            )
            self._done()
            return

        results = []
        try:
            for idx, (url, ean) in enumerate(zip(urls, eans), 1):
                if self._stop_scrape:
                    self.log("Stopped by user.", "warn")
                    break

                url = str(url).strip()
                if not url.startswith("http"):
                    self.log(f"[{idx}/{total}] Skipped (invalid URL): {url[:60]}", "warn")
                    continue

                self._set_status(f"Scraping {idx}/{total}…")
                pct = idx / total * 100
                self.progress_var.set(pct)

                result = scrape_product(self._driver, url, ean_from_excel=ean)
                results.append(result)

                if result["status"] == "ok":
                    self.log(
                        f"[{idx}/{total}] ✓ {result['product_name'][:50]} | "
                        f"EAN: {result['ean']} | "
                        f"Price: {result['discounted_price']} | "
                        f"Disc: {result['discount_pct']}%",
                        "info"
                    )
                else:
                    self.log(f"[{idx}/{total}] ✗ {url[:60]} — {result['error']}", "error")

                time.sleep(DELAY_BETWEEN_PRODUCTS)
        finally:
            try:
                if self._driver:
                    self._driver.quit()
            except Exception:
                pass
            self._driver = None
            self.log("Browser closed.", "head")

        if results:
            ok_count = sum(1 for r in results if r["status"] == "ok")
            self.log(f"══════════════════════════════════════", "head")
            self.log(f"Done! {ok_count}/{len(results)} products scraped OK.", "head")

            saved_msg_parts = []

            # ── Excel output (only if folder is set) ──
            if output_path:
                try:
                    out_file = save_results_to_excel(results, output_path, prev_prices)
                    self.log(f"Saved → {out_file}", "head")
                    saved_msg_parts.append(f"Excel:\n{out_file}")
                except Exception as e:
                    self.log(f"Error saving Excel: {e}", "error")

            # ── Google Sheets sync (only if URL is set) ──
            if sheet_url:
                self.log("Syncing to Google Sheets…", "head")
                ok = sync_to_google_sheets(
                    results, sheet_url, prev_prices, creds_path, log_fn=self.log
                )
                if ok:
                    saved_msg_parts.append(f"Google Sheet updated.")

            summary = (
                f"Scraping complete!\n\n"
                f"✓ {ok_count} products OK\n"
                f"✗ {len(results)-ok_count} errors\n\n"
                + ("\n\n".join(saved_msg_parts) if saved_msg_parts else "No output saved.")
            )
            messagebox.showinfo("Complete", summary)

        self._done()

    def _done(self):
        self._running = False
        self.run_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.progress_var.set(0)
        self._set_status("Ready")


if __name__ == "__main__":
    app = TrendyolScraperApp()
    app.mainloop()
